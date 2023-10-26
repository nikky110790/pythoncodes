# Check the number of not available plants in the inventory and list them

import openpyxl

# Loading the existing workbook
inventory_workbook = openpyxl.load_workbook(r"D:\My_Learning\Python\SolutionsAndReferences\InventoryOfPlantsExcel\Plants.xlsx")

# Assign a variable to the current sheet
inventory_sheet = inventory_workbook.active

# list to hold inventory
non_available_plants = []

print("Maximum row used in the worksheet is : " + str(inventory_sheet.max_row))

for record in range(2, inventory_sheet.max_row + 1):
    if inventory_sheet.cell(record,1).value != None:
        if inventory_sheet.cell(record,1).offset(0,7).value.lower() == "no":
            non_available_plants.append(inventory_sheet.cell(record,1).value)

# while inventory_sheet.cell(row_counter,1).value != None:

#     # Adding the non availalbe plants to the inventory
#     if inventory_sheet.cell(row_counter,1).offset(0,7).value.lower() == "no":
#         non_available_plants.append(inventory_sheet.cell(row_counter,1).value)

#     row_counter += 1

inventory_workbook.close()

# Printing each of the non available list
print("The list of plants that are not in stock are below:")
for plant in non_available_plants:
    print(plant)

