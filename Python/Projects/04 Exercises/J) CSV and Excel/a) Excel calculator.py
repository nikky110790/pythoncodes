# Excel calculator


import openpyxl

# Create workbook variable to create a excel workbook

new_workbook = openpyxl.Workbook()

# Accessing the current worksheet
current_sheet = new_workbook.active
current_sheet.title = "Calculator"

# Creating the ranges
current_sheet.cell(1,1).value = "First number ==>"
current_sheet.cell(2,1).value = "Second number ==>"

current_sheet.cell(1,2).value = 13
current_sheet.cell(2,2).value = 8
current_sheet.cell(3,2).value = "=B1 + B2"

new_workbook.save(r"D:\My_Learning\Python\SolutionsAndReferences\Excel Calculator\Calculator.xlsx")
new_workbook.close()

print("Completed!!!")